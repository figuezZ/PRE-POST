"""Pruebas de los reportes descargables de transferencia y servicio."""

from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader
import pytest

from src.analysis import analyze_service, analyze_transfer
from src.models import (
    BeamInput,
    ConcreteInput,
    DesignInput,
    LoadInput,
    PrestressInput,
    ProjectMetadata,
    RectangularSectionInput,
)
from src.reporting import (
    build_excel_report,
    build_pdf_report,
    build_report_tables,
    safe_report_stem,
)
from src.units import UnitSystem


@pytest.fixture
def analyzed_design():
    design = DesignInput(
        metadata=ProjectMetadata(
            "Prueba de servicio",
            "Equipo ICIV 1042",
            "2026-08-26",
            "0.5.0",
            "ACI 318-19 (provisional)",
        ),
        beam=BeamInput(10.0),
        section=RectangularSectionInput(0.4, 0.8),
        concrete=ConcreteInput(35e6, 45e6, 34e9, 25e3),
        loads=LoadInput(2_000.0, 5_000.0),
        prestress=PrestressInput(
            initial_force_n=1_000_000.0,
            eccentricity_m=-0.20,
            steel_area_m2=0.0007,
            steel_ultimate_strength_pa=1_860e6,
            time_dependent_loss_ratio=0.15,
        ),
    )
    return design, analyze_transfer(design), analyze_service(design)


def test_excel_report_is_openable_and_contains_service_results(analyzed_design):
    design, transfer, service = analyzed_design

    report = build_excel_report(design, transfer, service)
    workbook = load_workbook(BytesIO(report), data_only=True)
    sheet = workbook["Informe PRE-POST"]
    rows = {
        sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value
        for row in range(1, sheet.max_row + 1)
    }

    assert report.startswith(b"PK")
    assert sheet["A1"].value == "PRE-POST | Informe de resultados"
    assert sheet.sheet_view.showGridLines is False
    assert rows["Fuerza efectiva Pe"] == pytest.approx(850.0)
    assert rows["Carga uniforme total"] == pytest.approx(15.0)
    assert rows["Tension fibra superior"] == pytest.approx(
        service.stress.top_pa / 1e6
    )


def test_pdf_report_is_openable_and_contains_traceability(analyzed_design):
    design, transfer, service = analyzed_design

    report = build_pdf_report(design, transfer, service)
    reader = PdfReader(BytesIO(report))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert report.startswith(b"%PDF")
    assert len(reader.pages) == 2
    assert "PRE-POST | Informe de resultados" in text
    assert "Fuerza efectiva Pe" in text
    assert "Clase 3 USS (2026)" in text
    assert "Compresion negativa" in text


def test_uscs_report_converts_values_and_declares_units(analyzed_design):
    design, transfer, service = analyzed_design

    tables = build_report_tables(
        design, transfer, service, unit_system=UnitSystem.USCS
    )
    rows = {
        (table.title, row.label): (row.value, row.unit)
        for table in tables
        for row in table.rows
    }

    assert rows[("Identificacion", "Sistema de unidades")] == ("USCS", "")
    assert rows[("Datos de entrada", "Luz")][0] == pytest.approx(32.80839895)
    assert rows[("Datos de entrada", "Luz")][1] == "ft"
    assert rows[("Datos de entrada", "Fuerza inicial Pi")][0] == pytest.approx(
        224.8089431
    )
    assert rows[("Datos de entrada", "Fuerza inicial Pi")][1] == "kip"
    assert rows[("Resultados - servicio", "Momento en centro")][1] == "kip ft"
    assert rows[("Resultados - servicio", "Tension fibra superior")][1] == "psi"

    report = build_excel_report(
        design, transfer, service, unit_system=UnitSystem.USCS
    )
    workbook = load_workbook(BytesIO(report), data_only=True)
    sheet = workbook["Informe PRE-POST"]
    sheet_rows = {
        sheet.cell(row=row, column=1).value: (
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
        )
        for row in range(1, sheet.max_row + 1)
    }
    assert sheet_rows["Sistema de unidades"] == ("USCS", None)
    assert sheet_rows["Fuerza efectiva Pe"][1] == "kip"

    pdf_report = build_pdf_report(
        design, transfer, service, unit_system=UnitSystem.USCS
    )
    pdf_reader = PdfReader(BytesIO(pdf_report))
    pdf_text = "\n".join(
        page.extract_text() or "" for page in pdf_reader.pages
    )
    assert "Sistema de unidades" in pdf_text
    assert "USCS" in pdf_text
    assert "kip ft" in pdf_text
    assert "psi" in pdf_text


def test_report_filename_is_portable():
    assert safe_report_stem("Viga Ñuble / Etapa 1", "2026-08-26") == (
        "PRE_POST_Viga_Nuble_Etapa_1_2026-08-26"
    )
