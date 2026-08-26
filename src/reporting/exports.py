"""Exportacion tabulada de resultados calculados por el nucleo.

Este modulo no contiene formulas de ingenieria. Recibe resultados ya
calculados por ``src.analysis`` y solo convierte unidades para presentacion.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
import reportlab

from src.models import CaseAnalysisResult, DesignInput, ServiceAnalysisResult


PDF_FONT_REGULAR = "PrePostVera"
PDF_FONT_BOLD = "PrePostVeraBold"


def _register_pdf_fonts() -> None:
    """Incrusta fuentes portables incluidas con ReportLab."""

    registered = set(pdfmetrics.getRegisteredFontNames())
    font_directory = reportlab.__path__[0] + "/fonts"
    if PDF_FONT_REGULAR not in registered:
        pdfmetrics.registerFont(
            TTFont(PDF_FONT_REGULAR, f"{font_directory}/Vera.ttf")
        )
    if PDF_FONT_BOLD not in registered:
        pdfmetrics.registerFont(
            TTFont(PDF_FONT_BOLD, f"{font_directory}/VeraBd.ttf")
        )


@dataclass(frozen=True, slots=True)
class ReportRow:
    label: str
    value: str | float
    unit: str = ""
    decimals: int = 3


@dataclass(frozen=True, slots=True)
class ReportTable:
    title: str
    rows: tuple[ReportRow, ...]


def safe_report_stem(project_name: str, calculation_date: str) -> str:
    """Crea un nombre de archivo portable sin modificar el nombre mostrado."""

    normalized = unicodedata.normalize("NFKD", project_name)
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", ascii_name).strip("_")
    safe_date = re.sub(r"[^0-9-]+", "", calculation_date)
    return f"PRE_POST_{safe_name or 'proyecto'}_{safe_date or 'sin_fecha'}"


def build_report_tables(
    design: DesignInput,
    transfer: CaseAnalysisResult,
    service: ServiceAnalysisResult,
) -> tuple[ReportTable, ...]:
    """Organiza una instantanea comun para Excel y PDF.

    Las magnitudes provienen de ``analyze_transfer`` y ``analyze_service``.
    La referencia mecanica del flujo es la Clase 3 USS (2026), ecuaciones
    (25)-(28); aqui solo se transforman SI a las unidades visibles.
    """

    return (
        ReportTable(
            "Identificacion",
            (
                ReportRow("Proyecto", design.metadata.name),
                ReportRow("Autor", design.metadata.author),
                ReportRow("Fecha de calculo", design.metadata.calculation_date),
                ReportRow("Version PRE-POST", design.metadata.version),
                ReportRow("Norma declarada", design.metadata.standard),
            ),
        ),
        ReportTable(
            "Datos de entrada",
            (
                ReportRow("Luz", design.beam.span_m, "m"),
                ReportRow("Ancho b", design.section.width_m, "m"),
                ReportRow("Altura h", design.section.height_m, "m"),
                ReportRow(
                    "Peso especifico del hormigon",
                    design.concrete.unit_weight_n_m3 / 1e3,
                    "kN/m3",
                ),
                ReportRow(
                    "Resistencia f'ci",
                    design.concrete.compressive_strength_transfer_pa / 1e6,
                    "MPa",
                ),
                ReportRow(
                    "Resistencia f'c",
                    design.concrete.compressive_strength_service_pa / 1e6,
                    "MPa",
                ),
                ReportRow(
                    "Carga muerta adicional",
                    design.loads.superimposed_dead_load_n_m / 1e3,
                    "kN/m",
                ),
                ReportRow("Carga viva de servicio", design.loads.live_load_n_m / 1e3, "kN/m"),
                ReportRow("Fuerza inicial Pi", design.prestress.initial_force_n / 1e3, "kN"),
                ReportRow(
                    "Excentricidad e (+ arriba)",
                    design.prestress.eccentricity_m,
                    "m",
                ),
                ReportRow("Area de acero Ap", design.prestress.steel_area_m2 * 1e6, "mm2"),
                ReportRow(
                    "Perdida global declarada",
                    design.prestress.time_dependent_loss_ratio * 100.0,
                    "%",
                ),
            ),
        ),
        ReportTable(
            "Propiedades de la seccion",
            (
                ReportRow("Area", transfer.section.area_m2, "m2", 6),
                ReportRow(
                    "Centroide desde la base",
                    transfer.section.centroid_from_bottom_m,
                    "m",
                    6,
                ),
                ReportRow("Inercia", transfer.section.inertia_m4, "m4", 6),
                ReportRow(
                    "Modulo resistente superior",
                    transfer.section.section_modulus_top_m3,
                    "m3",
                    6,
                ),
                ReportRow(
                    "Modulo resistente inferior",
                    transfer.section.section_modulus_bottom_m3,
                    "m3",
                    6,
                ),
                ReportRow("Peso propio", transfer.section.self_weight_n_m / 1e3, "kN/m"),
            ),
        ),
        ReportTable(
            "Resultados - transferencia",
            (
                ReportRow(
                    "Carga uniforme",
                    transfer.transfer_uniform_load_n_m / 1e3,
                    "kN/m",
                ),
                ReportRow("Reaccion", transfer.transfer_reaction_n / 1e3, "kN"),
                ReportRow("Corte maximo", transfer.transfer_max_shear_n / 1e3, "kN"),
                ReportRow(
                    "Momento en centro",
                    transfer.transfer_midspan_moment_n_m / 1e3,
                    "kN m",
                ),
                ReportRow(
                    "Tension inicial del acero",
                    transfer.initial_steel_stress_pa / 1e6,
                    "MPa",
                ),
                ReportRow(
                    "Tension fibra superior",
                    transfer.transfer_stress.top_pa / 1e6,
                    "MPa",
                ),
                ReportRow(
                    "Tension fibra inferior",
                    transfer.transfer_stress.bottom_pa / 1e6,
                    "MPa",
                ),
            ),
        ),
        ReportTable(
            "Resultados - servicio",
            (
                ReportRow(
                    "Fuerza efectiva Pe",
                    service.effective_prestress_force_n / 1e3,
                    "kN",
                ),
                ReportRow(
                    "Tension efectiva del acero",
                    design.prestress.effective_stress_pa / 1e6,
                    "MPa",
                ),
                ReportRow(
                    "Carga uniforme total",
                    service.total_uniform_load_n_m / 1e3,
                    "kN/m",
                ),
                ReportRow("Reaccion", service.reaction_n / 1e3, "kN"),
                ReportRow("Corte maximo", service.max_shear_n / 1e3, "kN"),
                ReportRow(
                    "Momento en centro",
                    service.midspan_moment_n_m / 1e3,
                    "kN m",
                ),
                ReportRow("Tension fibra superior", service.stress.top_pa / 1e6, "MPa"),
                ReportRow("Tension fibra inferior", service.stress.bottom_pa / 1e6, "MPa"),
            ),
        ),
        ReportTable(
            "Trazabilidad y alcance",
            (
                ReportRow(
                    "Convencion de tensiones",
                    "Compresion negativa; traccion positiva",
                ),
                ReportRow(
                    "Convencion geometrica",
                    "e positiva hacia arriba; tendon inferior con e negativa",
                ),
                ReportRow(
                    "Referencia docente",
                    "Clase 3 USS (2026), ecuaciones (25)-(28)",
                ),
                ReportRow(
                    "Alcance",
                    "Resultados elasticos academicos sin verificacion normativa",
                ),
            ),
        ),
    )


def build_excel_report(
    design: DesignInput,
    transfer: CaseAnalysisResult,
    service: ServiceAnalysisResult,
) -> bytes:
    """Entrega un libro XLSX tabulado como instantanea de los resultados."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informe PRE-POST"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    navy = "17365D"
    blue = "1F4E78"
    pale_blue = "D9EAF7"
    pale_yellow = "FFF2CC"
    light_gray = "E7E6E6"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E2F3")

    sheet.merge_cells("A1:C1")
    title = sheet["A1"]
    title.value = "PRE-POST | Informe de resultados"
    title.font = Font(name="Aptos Display", size=18, bold=True, color=white)
    title.fill = PatternFill("solid", fgColor=navy)
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells("A2:C2")
    sheet["A2"] = design.metadata.name
    sheet["A2"].font = Font(name="Aptos", size=12, bold=True, color=navy)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24

    sheet.merge_cells("A3:C3")
    sheet["A3"] = (
        "Herramienta academica en desarrollo. No usar para construir obras reales."
    )
    sheet["A3"].fill = PatternFill("solid", fgColor=pale_yellow)
    sheet["A3"].font = Font(name="Aptos", size=10, italic=True, color="7F6000")
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 28

    row_index = 5
    for table in build_report_tables(design, transfer, service):
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        section_cell = sheet.cell(row=row_index, column=1, value=table.title)
        section_cell.fill = PatternFill("solid", fgColor=blue)
        section_cell.font = Font(name="Aptos", size=11, bold=True, color=white)
        section_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[row_index].height = 22
        row_index += 1

        headers = ("Magnitud", "Valor", "Unidad")
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column, value=header)
            cell.fill = PatternFill("solid", fgColor=pale_blue)
            cell.font = Font(name="Aptos", size=10, bold=True, color=navy)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(bottom=thin_gray)
        row_index += 1

        for offset, report_row in enumerate(table.rows):
            background = light_gray if offset % 2 else white
            label_cell = sheet.cell(row=row_index, column=1, value=report_row.label)
            value_cell = sheet.cell(row=row_index, column=2, value=report_row.value)
            unit_cell = sheet.cell(row=row_index, column=3, value=report_row.unit)
            for cell in (label_cell, value_cell, unit_cell):
                cell.fill = PatternFill("solid", fgColor=background)
                cell.font = Font(name="Aptos", size=10, color="1F1F1F")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin_gray)
            if isinstance(report_row.value, (int, float)):
                value_cell.number_format = f"#,##0.{''.join('0' for _ in range(report_row.decimals))}"
                value_cell.alignment = Alignment(horizontal="right", vertical="top")
            row_index += 1
        row_index += 1

    widths = (42, 34, 14)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.print_title_rows = "1:3"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "PRE-POST - pagina &P de &N"
    workbook.properties.title = "PRE-POST | Informe de resultados"
    workbook.properties.subject = "Transferencia y servicio elastico"
    workbook.properties.creator = design.metadata.author

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_value(row: ReportRow) -> str:
    if isinstance(row.value, (int, float)):
        return f"{row.value:,.{row.decimals}f}"
    return str(row.value)


def _draw_pdf_footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setFont(PDF_FONT_REGULAR, 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawString(18 * mm, 11 * mm, "PRE-POST | Uso academico")
    canvas.drawRightString(192 * mm, 11 * mm, f"Pagina {document.page}")
    canvas.restoreState()


def build_pdf_report(
    design: DesignInput,
    transfer: CaseAnalysisResult,
    service: ServiceAnalysisResult,
) -> bytes:
    """Entrega un PDF tabulado como instantanea de los resultados."""

    _register_pdf_fonts()
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title="PRE-POST | Informe de resultados",
        author=design.metadata.author,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PrePostTitle",
        parent=styles["Title"],
        fontName=PDF_FONT_BOLD,
        fontSize=19,
        leading=23,
        textColor=colors.HexColor("#17365D"),
        alignment=TA_CENTER,
        spaceAfter=4 * mm,
    )
    project_style = ParagraphStyle(
        "PrePostProject",
        parent=styles["Heading2"],
        fontName=PDF_FONT_BOLD,
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#1F4E78"),
        alignment=TA_CENTER,
        spaceAfter=4 * mm,
    )
    body_style = ParagraphStyle(
        "PrePostBody",
        parent=styles["BodyText"],
        fontName=PDF_FONT_REGULAR,
        fontSize=8.5,
        leading=11,
    )
    section_style = ParagraphStyle(
        "PrePostSection",
        parent=body_style,
        fontName=PDF_FONT_BOLD,
        textColor=colors.white,
    )
    header_style = ParagraphStyle(
        "PrePostHeader",
        parent=body_style,
        fontName=PDF_FONT_BOLD,
        textColor=colors.HexColor("#17365D"),
    )

    story = [
        Paragraph("PRE-POST | Informe de resultados", title_style),
        Paragraph(design.metadata.name, project_style),
    ]
    warning = Table(
        [[Paragraph(
            "Herramienta academica en desarrollo. No usar para construir obras reales.",
            body_style,
        )]],
        colWidths=[174 * mm],
    )
    warning.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF2CC")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#7F6000")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6B656")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend((warning, Spacer(1, 5 * mm)))

    for index, report_table in enumerate(build_report_tables(design, transfer, service)):
        if index == 3:
            story.append(PageBreak())
        data = [
            [Paragraph(report_table.title, section_style), "", ""],
            [
                Paragraph("Magnitud", header_style),
                Paragraph("Valor", header_style),
                Paragraph("Unidad", header_style),
            ],
        ]
        for row in report_table.rows:
            data.append(
                [
                    Paragraph(row.label, body_style),
                    Paragraph(_pdf_value(row), body_style),
                    Paragraph(row.unit, body_style),
                ]
            )
        table = LongTable(data, colWidths=[91 * mm, 61 * mm, 22 * mm], repeatRows=2)
        commands = [
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D9EAF7")),
            ("GRID", (0, 1), (-1, -1), 0.25, colors.HexColor("#D9E2F3")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ALIGN", (1, 2), (1, -1), "RIGHT"),
        ]
        for row_number in range(2, len(data)):
            if row_number % 2:
                commands.append(
                    ("BACKGROUND", (0, row_number), (-1, row_number), colors.HexColor("#F4F6F8"))
                )
        table.setStyle(TableStyle(commands))
        story.extend((table, Spacer(1, 4 * mm)))

    document.build(story, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    return output.getvalue()
